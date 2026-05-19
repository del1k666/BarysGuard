import os
import re
import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel, QLineEdit,
    QPushButton, QTextEdit, QFileDialog, QCheckBox
)
from constants import BUILTIN_YARA_RULES
from ui.dashboard_tab import DashboardTab
from core.i18n import t


class ReportTab(QWidget):
    def __init__(self):
        super().__init__()
        self._build()

    def _build(self):
        lay = QVBoxLayout(self); lay.setSpacing(10); lay.setContentsMargins(16,16,16,16)

        self._grp_params = QGroupBox(t("report_params_grp"))
        gl = QVBoxLayout(self._grp_params); gl.setSpacing(8)

        row1 = QHBoxLayout()
        self._lbl_analyst = QLabel(t("report_analyst_lbl")); row1.addWidget(self._lbl_analyst)
        self.analyst = QLineEdit(t("report_analyst_def")); row1.addWidget(self.analyst)
        self._lbl_org = QLabel(t("report_org_lbl")); row1.addWidget(self._lbl_org)
        self.org = QLineEdit("BarysGuard"); row1.addWidget(self.org)
        gl.addLayout(row1)

        row2 = QHBoxLayout()
        self._lbl_title = QLabel(t("report_title_lbl")); row2.addWidget(self._lbl_title)
        self.title_inp = QLineEdit(t("report_title_def"))
        row2.addWidget(self.title_inp); gl.addLayout(row2)

        chk_row = QHBoxLayout()
        self.chk_stats     = QCheckBox(t("report_chk_stats"));  self.chk_stats.setChecked(True)
        self.chk_events    = QCheckBox(t("report_chk_events")); self.chk_events.setChecked(True)
        self.chk_yara_rules= QCheckBox(t("report_chk_yara"));   self.chk_yara_rules.setChecked(True)
        self.chk_net       = QCheckBox(t("report_chk_net"));    self.chk_net.setChecked(True)
        for c in [self.chk_stats, self.chk_events, self.chk_yara_rules, self.chk_net]:
            chk_row.addWidget(c)
        chk_row.addStretch()
        gl.addLayout(chk_row)
        lay.addWidget(self._grp_params)

        btn_row = QHBoxLayout()
        self._btn_preview = QPushButton(t("report_preview_btn"))
        self._btn_preview.setObjectName("secondaryBtn"); self._btn_preview.setFixedHeight(36)
        self._btn_preview.clicked.connect(self._preview); btn_row.addWidget(self._btn_preview)

        self._btn_html = QPushButton(t("report_html_btn")); self._btn_html.setFixedHeight(36)
        self._btn_html.clicked.connect(self._export_html); btn_row.addWidget(self._btn_html)

        self._btn_xl = QPushButton(t("report_excel_btn"))
        self._btn_xl.setObjectName("secondaryBtn"); self._btn_xl.setFixedHeight(36)
        self._btn_xl.clicked.connect(self._export_excel); btn_row.addWidget(self._btn_xl)

        self._btn_txt = QPushButton(t("report_txt_btn"))
        self._btn_txt.setObjectName("secondaryBtn"); self._btn_txt.setFixedHeight(36)
        self._btn_txt.clicked.connect(self._export_txt); btn_row.addWidget(self._btn_txt)
        lay.addLayout(btn_row)

        self.preview = QTextEdit(); self.preview.setReadOnly(True)
        self.preview.setPlaceholderText(t("report_placeholder"))
        lay.addWidget(self.preview)

    def retranslate(self, _lang: str = ""):
        self._grp_params.setTitle(t("report_params_grp"))
        self._lbl_analyst.setText(t("report_analyst_lbl"))
        self._lbl_org.setText(t("report_org_lbl"))
        self._lbl_title.setText(t("report_title_lbl"))
        self.chk_stats.setText(t("report_chk_stats"))
        self.chk_events.setText(t("report_chk_events"))
        self.chk_yara_rules.setText(t("report_chk_yara"))
        self.chk_net.setText(t("report_chk_net"))
        self._btn_preview.setText(t("report_preview_btn"))
        self._btn_html.setText(t("report_html_btn"))
        self._btn_xl.setText(t("report_excel_btn"))
        self._btn_txt.setText(t("report_txt_btn"))
        self.preview.setPlaceholderText(t("report_placeholder"))

    def _collect_data(self):
        """Собирает все реальные данные из DashboardTab.stats"""
        s = DashboardTab.stats
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        events = s.get("recent", [])

        # Разбиваем события по типам
        hash_events  = [e for e in events if e.get("type") == "HASH"]
        yara_events  = [e for e in events if e.get("type") == "YARA"]
        net_events   = [e for e in events if e.get("type") == "NET"]
        ioc_events   = [e for e in events if e.get("type") == "IOC"]
        quar_events    = [e for e in events if e.get("type") == "QUARANTINE"]
        remote_events  = [e for e in events if e.get("host")]

        # Severity counts из YARA событий
        sev_counts = {"Critical":0,"High":0,"Medium":0,"Low":0}
        _level_map = {"critical":"Critical","high":"High","medium":"Medium","low":"Low"}
        for e in yara_events:
            sv = e.get("severity","")
            if not sv or sv.upper() == e.get("type","").upper():
                sv = _level_map.get(e.get("level","").lower(), "")
            sv = sv.capitalize()
            if sv in sev_counts:
                sev_counts[sv] += 1

        vt_hits = [r for r in DashboardTab.remote_hash_vt_results
                   if r.get("status") in ("MALICIOUS", "SUSPICIOUS")]

        return {
            "title":    self.title_inp.text(),
            "analyst":  self.analyst.text(),
            "org":      self.org.text(),
            "date":     now,
            "stats":    s,
            "hash_events":  hash_events,
            "yara_events":  yara_events,
            "net_events":   net_events,
            "ioc_events":   ioc_events,
            "quar_events":    quar_events,
            "remote_events":  remote_events,
            "sev_counts":     sev_counts,
            "vt_hits":        vt_hits,
            "total_events": len(events),
        }

    def _preview(self):
        d = self._collect_data()
        s = d["stats"]
        lines = []
        lines.append("=" * 65)
        lines.append(f"  {d['title'].upper()}")
        lines.append("=" * 65)
        lines.append(f"  Аналитик     : {d['analyst']}")
        lines.append(f"  Организация  : {d['org']}")
        lines.append(f"  Дата         : {d['date']}")
        lines.append("")

        if self.chk_stats.isChecked():
            lines.append("СТАТИСТИКА СЕССИИ")
            lines.append("-" * 40)
            lines.append(f"  Hash Lookup проверок  : {s['hash_lookups']}")
            lines.append(f"  Malicious файлов      : {s['malicious']}")
            lines.append(f"  Чистых файлов         : {s['clean']}")
            lines.append(f"  IOC сборов            : {s['ioc_runs']}")
            lines.append(f"  Подозрительных проц.  : {s['suspicious_procs']}")
            lines.append(f"  YARA сканирований     : {s['yara_scans']}")
            lines.append(f"  YARA совпадений       : {s['yara_hits']}")
            lines.append(f"  Проверок IP/домен     : {s['net_checks']}")
            lines.append(f"  Высокорисковых IP     : {s['high_risk_ips']}")
            lines.append("")

        sev = d["sev_counts"]
        if any(sev.values()):
            lines.append("РАСПРЕДЕЛЕНИЕ УГРОЗ")
            lines.append("-" * 40)
            for level, count in sev.items():
                bar = "#" * count + "." * max(0, 20-count)
                lines.append(f"  {level:<10}: [{bar}] {count}")
            lines.append("")

        if self.chk_events.isChecked() and d["yara_events"]:
            lines.append(f"YARA ДЕТЕКТЫ ({len(d['yara_events'])})")
            lines.append("-" * 40)
            for e in d["yara_events"]:
                lines.append(f"  [{e.get('severity','?').upper()[:4]}] {e.get('msg','')}")
            lines.append("")

        if self.chk_events.isChecked() and d["hash_events"]:
            lines.append(f"HASH LOOKUP ({len(d['hash_events'])})")
            lines.append("-" * 40)
            for e in d["hash_events"]:
                lines.append(f"  {e.get('msg','')}")
            lines.append("")

        if d.get("vt_hits"):
            lines.append(f"VT REMOTE SCAN — УГРОЗЫ ({len(d['vt_hits'])})")
            lines.append("-" * 40)
            for r in d["vt_hits"]:
                lines.append(
                    f"  [{r.get('status','?'):<10}] {r.get('host','')} — {r.get('file','')}  "
                    f"({r.get('mal',0)}/{r.get('total',0)} engines)")
            lines.append("")

        if self.chk_net.isChecked() and d["net_events"]:
            lines.append(f"NETWORK INTEL ({len(d['net_events'])})")
            lines.append("-" * 40)
            for e in d["net_events"]:
                lines.append(f"  {e.get('target','')} — {e.get('msg','')}")
            lines.append("")

        if d["ioc_events"]:
            lines.append(f"IOC COLLECTION ({len(d['ioc_events'])})")
            lines.append("-" * 40)
            for e in d["ioc_events"]:
                lines.append(f"  {e.get('time','')} {e.get('msg','')}")
            lines.append("")

        if d["quar_events"]:
            lines.append(f"QUARANTINE ({len(d['quar_events'])})")
            lines.append("-" * 40)
            for e in d["quar_events"]:
                lines.append(f"  {e.get('msg','')}")
            lines.append("")

        if d.get("remote_events"):
            by_host: dict = {}
            for e in d["remote_events"]:
                by_host.setdefault(e.get("host", "Unknown"), []).append(e)
            lines.append(f"УДАЛЁННЫЕ СКАНЫ ({len(d['remote_events'])})")
            lines.append("-" * 40)
            for host_name, evts in by_host.items():
                lines.append(f"  Хост: {host_name}")
                for e in evts:
                    lines.append(
                        f"    [{e.get('type','?')}] {e.get('time','')}  {e.get('msg','')}"
                    )
            lines.append("")

        if self.chk_yara_rules.isChecked():
            lines.append("ВСТРОЕННЫЕ YARA ПРАВИЛА")
            lines.append("-" * 40)
            for name, text in BUILTIN_YARA_RULES.items():
                m = re.search(r'severity\s*=\s*"(\w+)"', text)
                sev_r = m.group(1).upper() if m else "INFO"
                lines.append(f"  [{sev_r:<8}] {name}")
            lines.append("")

        if d["total_events"] == 0:
            lines.append("  Нет данных — запустите сканирования перед генерацией отчёта.")

        lines.append("=" * 65)
        self.preview.setText("\n".join(lines))

    def _export_html(self):
        d = self._collect_data()
        s = d["stats"]

        # Severity bar helper
        def sev_bar(count, total, color):
            pct = int(count / max(total, 1) * 100)
            return (f'<div style="display:flex;align-items:center;gap:8px;margin:4px 0;">'
                    f'<div style="width:{pct*2}px;height:12px;background:{color};border-radius:3px;min-width:2px"></div>'
                    f'<span style="color:{color};font-weight:bold">{count}</span></div>')

        sev = d["sev_counts"]
        total_sev = max(sum(sev.values()), 1)
        sev_html = ""
        for level, col in [("Critical","#f85149"),("High","#d29922"),("Medium","#58a6ff"),("Low","#3fb950")]:
            sev_html += f'<tr><td style="color:{col};font-weight:bold">{level}</td><td>{sev_bar(sev[level],total_sev,col)}</td><td style="color:{col}">{sev[level]}</td></tr>\n'

        # Events table
        yara_rows = ""
        for e in d["yara_events"]:
            sv = e.get("severity","info")
            col = {"Critical":"#f85149","High":"#d29922","Medium":"#58a6ff","Low":"#3fb950"}.get(sv,"#8b949e")
            yara_rows += f'<tr><td>{e.get("time","")}</td><td style="color:{col}">{sv}</td><td>{e.get("msg","")}</td></tr>\n'

        hash_rows = ""
        for e in d["hash_events"]:
            lvl = e.get("level","info")
            col = {"critical":"#f85149","high":"#d29922","ok":"#3fb950"}.get(lvl,"#8b949e")
            hash_rows += f'<tr><td>{e.get("time","")}</td><td style="color:{col}">{e.get("msg","")}</td></tr>\n'

        net_rows = ""
        for e in d["net_events"]:
            lvl = e.get("level","info")
            col = {"critical":"#f85149","high":"#d29922","ok":"#3fb950"}.get(lvl,"#8b949e")
            net_rows += f'<tr><td>{e.get("target","")}</td><td style="color:{col}">{e.get("msg","")}</td></tr>\n'

        html = f"""<!DOCTYPE html>
<html lang="ru"><head><meta charset="UTF-8"><title>{d['title']}</title>
<style>
body{{background:#0d1117;color:#e6edf3;font-family:'Segoe UI',sans-serif;padding:40px;max-width:1100px;margin:0 auto;line-height:1.6;}}
h1{{color:#58a6ff;border-bottom:2px solid #21262d;padding-bottom:12px;font-size:24px;}}
h2{{color:#3fb950;margin-top:32px;font-size:16px;text-transform:uppercase;letter-spacing:1px;}}
.meta-box{{background:#161b22;border:1px solid #21262d;border-radius:8px;padding:20px;margin:20px 0;display:grid;grid-template-columns:1fr 1fr;gap:8px;}}
.meta-row{{display:flex;gap:12px;}}
.meta-key{{color:#6e7681;font-size:12px;min-width:120px;}}
.meta-val{{color:#e6edf3;font-size:12px;}}
.section{{background:#161b22;border:1px solid #21262d;border-radius:8px;padding:20px;margin:20px 0;}}
.stats-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin:12px 0;}}
.stat-card{{background:#0d1117;border:1px solid #21262d;border-radius:6px;padding:14px;}}
.stat-num{{font-size:32px;font-weight:bold;margin-bottom:4px;}}
.stat-lbl{{color:#6e7681;font-size:11px;text-transform:uppercase;letter-spacing:1px;}}
table{{width:100%;border-collapse:collapse;margin-top:10px;font-size:13px;}}
th{{background:#21262d;color:#6e7681;padding:8px 10px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:0.5px;}}
td{{padding:7px 10px;border-bottom:1px solid #161b22;}}
tr:hover td{{background:#1c2128;}}
.critical{{color:#f85149;}} .high{{color:#d29922;}} .medium{{color:#58a6ff;}} .low{{color:#3fb950;}}
.empty{{color:#484f58;font-style:italic;padding:20px;text-align:center;}}
.footer{{color:#484f58;font-size:11px;text-align:center;margin-top:40px;border-top:1px solid #21262d;padding-top:16px;}}
</style></head><body>
<h1>{d['title']}</h1>
<div class="meta-box">
<div class="meta-row"><span class="meta-key">Аналитик</span><span class="meta-val">{d['analyst']}</span></div>
<div class="meta-row"><span class="meta-key">Организация</span><span class="meta-val">{d['org']}</span></div>
<div class="meta-row"><span class="meta-key">Дата формирования</span><span class="meta-val">{d['date']}</span></div>
<div class="meta-row"><span class="meta-key">Инструмент</span><span class="meta-val">BarysGuard</span></div>
</div>

<h2>Статистика сессии</h2>
<div class="section">
<div class="stats-grid">
<div class="stat-card"><div class="stat-num" style="color:#58a6ff">{s['hash_lookups']}</div><div class="stat-lbl">Hash Lookups</div></div>
<div class="stat-card"><div class="stat-num" style="color:#f85149">{s['malicious']}</div><div class="stat-lbl">Malicious</div></div>
<div class="stat-card"><div class="stat-num" style="color:#d29922">{s['yara_hits']}</div><div class="stat-lbl">YARA Hits</div></div>
<div class="stat-card"><div class="stat-num" style="color:#3fb950">{s['ioc_runs']}</div><div class="stat-lbl">IOC Runs</div></div>
<div class="stat-card"><div class="stat-num" style="color:#3fb950">{s['net_checks']}</div><div class="stat-lbl">Net Checks</div></div>
<div class="stat-card"><div class="stat-num" style="color:#f85149">{s['high_risk_ips']}</div><div class="stat-lbl">High Risk IPs</div></div>
</div>
</div>

<h2>Распределение угроз</h2>
<div class="section"><table>
<tr><th>Уровень</th><th>График</th><th>Количество</th></tr>
{sev_html}
</table></div>"""

        if d["yara_events"]:
            html += f"""<h2>YARA Детекты ({len(d["yara_events"])})</h2>
<div class="section"><table>
<tr><th>Время</th><th>Severity</th><th>Событие</th></tr>
{yara_rows}</table></div>"""

        if d["hash_events"]:
            html += f"""<h2>Hash Lookup ({len(d["hash_events"])})</h2>
<div class="section"><table>
<tr><th>Время</th><th>Результат</th></tr>
{hash_rows}</table></div>"""

        if d.get("vt_hits"):
            vt_rows = ""
            for r in d["vt_hits"]:
                st  = r.get("status","?")
                col = "#f85149" if st == "MALICIOUS" else "#d29922"
                vt_rows += (
                    f'<tr><td style="color:#58a6ff">{r.get("host","")}</td>'
                    f'<td style="color:{col};font-weight:bold">{st}</td>'
                    f'<td>{r.get("file","")}</td>'
                    f'<td style="color:{col}">{r.get("mal",0)}/{r.get("total",0)}</td></tr>\n'
                )
            html += (
                f'<h2>VT Remote Hash Scan — Угрозы ({len(d["vt_hits"])})</h2>'
                f'<div class="section"><table>'
                f'<tr><th>Хост</th><th>Статус</th><th>Файл</th><th>Engines</th></tr>'
                f'{vt_rows}</table></div>'
            )

        if d["net_events"]:
            html += f"""<h2>Network Intelligence ({len(d["net_events"])})</h2>
<div class="section"><table>
<tr><th>Цель</th><th>Результат</th></tr>
{net_rows}</table></div>"""

        if d.get("remote_events"):
            by_host: dict = {}
            for e in d["remote_events"]:
                by_host.setdefault(e.get("host", "Unknown"), []).append(e)
            rem_colors = {"YARA": "#58a6ff", "IOC": "#d29922",
                          "MEMORY": "#a371f7", "HASH": "#8b949e"}
            remote_rows = ""
            for host_name, evts in by_host.items():
                for e in evts:
                    typ = e.get("type", "?")
                    col = rem_colors.get(typ, "#8b949e")
                    remote_rows += (
                        f'<tr><td>{e.get("time","")}</td>'
                        f'<td style="color:#58a6ff">{host_name}</td>'
                        f'<td style="color:{col}">{typ}</td>'
                        f'<td>{e.get("msg","")}</td></tr>\n'
                    )
            html += (
                f'<h2>Удалённые сканы ({len(d["remote_events"])})</h2>'
                f'<div class="section"><table>'
                f'<tr><th>Время</th><th>Хост</th><th>Тип</th><th>Правило / Файл</th></tr>'
                f'{remote_rows}</table></div>'
            )

        if self.chk_yara_rules.isChecked():
            yara_rule_rows = ""
            for name, text in BUILTIN_YARA_RULES.items():
                m = re.search(r'severity\s*=\s*"(\w+)"', text)
                d2 = re.search(r'description\s*=\s*"([^"]+)"', text)
                sev_r = m.group(1) if m else "info"
                desc  = d2.group(1) if d2 else ""
                yara_rule_rows += f'<tr><td>{name}</td><td class="{sev_r}">{sev_r.upper()}</td><td>{desc}</td></tr>\n'
            html += f"""<h2>YARA База правил</h2>
<div class="section"><table>
<tr><th>Правило</th><th>Severity</th><th>Описание</th></tr>
{yara_rule_rows}</table></div>"""

        html += f'''<div class="footer">Сформировано: {d['date']} | BarysGuard | {d['org']}</div>
</body></html>'''

        path, _ = QFileDialog.getSaveFileName(self, "Сохранить HTML отчёт", "report.html", "HTML (*.html)")
        if path:
            with open(path, "w", encoding="utf-8") as f:
                f.write(html)
            self.preview.setText(f"Отчёт сохранён: {path}\n\nОткрываю в браузере...")
            try: os.startfile(path)
            except: pass

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            self.preview.setText("openpyxl не установлен. Запусти: pip install openpyxl")
            return

        d = self._collect_data()
        s = d["stats"]
        wb = openpyxl.Workbook()

        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1F3864")

        def _sheet(ws, headers, rows):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = hdr_font
                cell.fill = hdr_fill
            for row in rows:
                ws.append(row)
            for col in ws.columns:
                width = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(width + 4, 60)

        ws = wb.active
        ws.title = "Статистика"
        _sheet(ws, ["Метрика", "Значение"], [
            ["Hash Lookups",     s["hash_lookups"]],
            ["Malicious",        s["malicious"]],
            ["Clean",            s["clean"]],
            ["IOC Runs",         s["ioc_runs"]],
            ["Suspicious Procs", s["suspicious_procs"]],
            ["YARA Scans",       s["yara_scans"]],
            ["YARA Hits",        s["yara_hits"]],
            ["Net Checks",       s["net_checks"]],
            ["High Risk IPs",    s["high_risk_ips"]],
        ])

        ws2 = wb.create_sheet("YARA")
        _sheet(ws2, ["Время", "Severity", "Событие"],
               [[e.get("time", ""), e.get("severity", ""), e.get("msg", "")]
                for e in d["yara_events"]])

        ws3 = wb.create_sheet("Hash Lookup")
        _sheet(ws3, ["Время", "Результат"],
               [[e.get("time", ""), e.get("msg", "")] for e in d["hash_events"]])

        ws4 = wb.create_sheet("IOC")
        _sheet(ws4, ["Время", "Тип", "Детали"],
               [[e.get("time", ""), e.get("type", ""), e.get("msg", "")]
                for e in d["ioc_events"]])

        ws5 = wb.create_sheet("Network")
        _sheet(ws5, ["Цель", "Результат"],
               [[e.get("target", ""), e.get("msg", "")] for e in d["net_events"]])

        ws6 = wb.create_sheet("Remote Scans")
        _sheet(ws6, ["Время", "Хост", "Тип", "Правило/Файл"],
               [[e.get("time", ""), e.get("host", ""), e.get("type", ""), e.get("msg", "")]
                for e in d["remote_events"]])

        ws7 = wb.create_sheet("VT Remote")
        _sheet(ws7, ["Хост", "Файл", "SHA256", "Статус", "Malicious", "Total"],
               [[r.get("host",""), r.get("file",""), r.get("sha256",""),
                 r.get("status",""), r.get("mal",0), r.get("total",0)]
                for r in d.get("vt_hits", [])])

        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить Excel отчёт", "report.xlsx", "Excel (*.xlsx)")
        if path:
            wb.save(path)
            self.preview.setText(f"Excel отчёт сохранён: {path}")

    def _export_txt(self):
        self._preview()
        txt = self.preview.toPlainText()
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить TXT отчёт", "report.txt", "Text (*.txt)")
        if path:
            with open(path, "w", encoding="utf-8") as f:
                f.write(txt)
            self.preview.append(f"\nСохранено: {path}")
